from django.shortcuts import render, redirect , get_object_or_404 
from .models import Cliente, Orden, Equipo, Tecnico
from django.views import generic
from django.core.urlresolvers import reverse_lazy
from django.core.urlresolvers import reverse
from django import forms
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook

from django.views.generic.edit import CreateView, UpdateView, DeleteView

# Create your views here.


def redirect_by_id(request):
    #if request.method == "POST":
    if  request.POST['id_orden']:
        id = request.POST['id_orden']
        return redirect('index:Detalle-Ordenes', pk=id)
    return redirect('index:home-page')






def ImprimirOrden(request,pk):
    import datetime
    now = datetime.datetime.now().strftime('%H:%M:%S') 
    orden=Orden.objects.get(pk = pk)
    context={'orden':orden, 'hora':now} 
    template = 'index/imprimir_orden.html'
    return render(request, template,context )


def ExcelExport (request):
    from django.db.models import Sum
    f_ini = request.POST['fecha_inicial']
    f_fin = request.POST['fecha_final']
    
    #configuracion libro
    wb = Workbook()
    dest_filename = 'file.xlsx'
    ws = wb.worksheets[0]
    
    if f_ini and f_fin:
        a = Orden.objects.filter(Fecha_creacion__range=(f_ini, f_fin))
    else:
        a = Orden.objects.all()

    suma= a.aggregate(Sum('Costo_reparacion'))
    num=0
    ws.cell(row=1, column=1, value="desde "+ str(f_ini))
    ws.cell(row=1, column=3, value="hasta "+ str(f_fin))

    ft = Font(bold=True)
    ws.cell(row=3, column=1, value="Id orden").font=ft
    ws.cell(row=3, column=2, value="Nombre Cli").font=ft
    ws.cell(row=3, column=3, value="Telefono").font=ft
    ws.cell(row=3, column=4, value="Equipo").font=ft
    ws.cell(row=3, column=5, value="Reparación").font=ft
    ws.cell(row=3, column=6, value="Costo").font=ft
    ws.cell(row=3, column=7, value="Estado").font=ft
    
    for orden in a:
        num= num + 1
        valor= orden.Costo_reparacion
        
        ws.cell(row=num+4, column=1, value=str(orden.pk))
        ws.cell(row=num+4, column=2, value=str(orden.Cliente.Nombre))
        ws.cell(row=num+4, column=3, value=str(orden.Cliente.Telefono1))
        ws.cell(row=num+4, column=4, value=str(orden.Equipo))
        ws.cell(row=num+4, column=5, value=str(orden.Informe_tecnico))
        ws.cell(row=num+4, column=6, value=str(orden.Costo_reparacion))
        ws.cell(row=num+4, column=7, value=str(orden.Estado))
        

    ws.cell(row=num+4, column=9, value="Total: "+str(suma['Costo_reparacion__sum']))

    return HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')


class ListaOrdenes(generic.ListView):
    #definir el template que utilizará y pasará el contexto
    template_name = 'index/ordenes.html'
    #se define el nombre del contexto
    context_object_name='all_ordenes'

    # @Override devuelve los objetos que serán renderizados
    def get_queryset(self):
        return Orden.objects.all().order_by('-Fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super(ListaOrdenes, self).get_context_data(**kwargs)
        context['pendientes'] = Orden.objects.filter(Estado='PROCESANDO')
        context['espera'] = Orden.objects.filter(Estado='ESPERA')
        return context

#no es factible porque la clase a la que hereda devuelve dos contextos
#y uno no se utiliza en esta, pero valio la pena la practica
class TodasOrdenes(ListaOrdenes):
    template_name = 'index/lista_ordenes.html'


class DetalleOrdenes(generic.DetailView):
    model=Orden
    template_name = 'index/detalle_orden.html'


#no se utiliza (didáctico)
class CrearOrden (CreateView):
    model= Orden
    #fields Im gonna let users to fill
    fields =['Fecha_entrega','Estado_inicial','Falla', 'Costo_reparacion', 'Costo_revision', 'Notas',
    'Fecha_ofrecida','Accesorios','Limite_garantia','Informe_tecnico','Tecnico','Equipo','Cliente','Estado']
    
    def get_initial(self):
        if self.request.GET.get('txtcliente'):
            a=self.request.GET.get('txtcliente')
            return {
            'Tecnico':4, 'Cliente':a,
            }
        else:
            #C = get_object_or_404(Recipe, slug=self.kwargs.get('slug'))
            return {
            'Tecnico':4,
            }


def agregarClienteOrden(request):
    #if request.method == "POST":
    if  request.POST.get('txtcliente'):
        cedula = request.POST.get('txtcliente')
        try:
            user = Cliente.objects.get(Cedula=cedula)
            return redirect('index:Add-Orden2', user.pk)
        except Cliente.DoesNotExist:
            return redirect('/#no')

    return redirect('index:Add-Orden2',0,)



class OrdenForm(forms.ModelForm):
    class Meta:
        model=Orden
        fields =['Fecha_entrega','Estado_inicial','Falla', 'Costo_reparacion', 'Costo_revision', 'Notas',
        'Fecha_ofrecida','Accesorios','Limite_garantia','Informe_tecnico','Tecnico','Equipo','Cliente','Estado']

    def __init__(self, *args, **kwargs): 
        user = kwargs.pop('user', None) # pop the 'user' from kwargs dictionary      
        super(OrdenForm, self).__init__(*args, **kwargs)
        self.fields['Equipo'] = forms.ModelChoiceField(queryset=Equipo.objects.filter(Cliente=user)) 


class CrearOrden2 (CreateView):
    model= Orden
    form_class = OrdenForm
    #fields Im gonna let users to fill
   # fields =['Fecha_entrega','Estado_inicial','Falla', 'Costo_reparacion', 'Costo_revision', 'Notas',
   # 'Fecha_ofrecida','Accesorios','Limite_garantia','Informe_tecnico','Tecnico','Equipo','Cliente','Estado']
    
    def get_initial(self):
        if self.kwargs['cliente'] != 0:
            a=self.kwargs['cliente']
            return {
            'Tecnico':4, 'Cliente':a,
            }
        else:
            return {
            'Tecnico':4,
            }

    def get_form_kwargs(self):
        kwargs = super(CrearOrden2, self).get_form_kwargs()
        kwargs['user'] = self.kwargs['cliente'] # pass the 'user' in kwargs
        return kwargs 


    

class EditarOrden (UpdateView):
    model= Orden
    #fields Im gonna let users to edit
    fields =['Falla', 'Costo_reparacion', 'Notas',
    'Informe_tecnico','Estado']

class EliminarOrden (DeleteView):
    model= Orden
    success_url= reverse_lazy("index:home-page")

    #sobreescribe metodo get para q actue como post y no se necesite form d confirmacion
    #def get(self, request, *args, **kwargs):
    #    return self.post(request, *args, **kwargs)


#...............Client Views

class ListaCliente(generic.ListView):
    #definir el template que utilizará y pasará el contexto
    template_name = 'index/clientes.html'
    #se define el nombre del contexto
    context_object_name='all_clientes'

    # @Override devuelve los objetos que serán renderizados
    def get_queryset(self):
        return Cliente.objects.all()

    def get_context_data(self, **kwargs):
        context = super(ListaCliente, self).get_context_data(**kwargs)
        context['ultimos'] = Cliente.objects.order_by('-Fecha_ingreso')[:5]
        return context



class DetalleCliente(generic.ListView):
    #definir el template que utilizará y pasará el contexto
    template_name = 'index/detalle_cliente.html'
    #se define el nombre del contexto
    context_object_name='all_equipos'

    # @Override devuelve los objetos que serán renderizados
    def get_queryset(self):
        return Equipo.objects.filter(Cliente=self.kwargs['pk'])#pk -> parametro pasado por url

    def get_context_data(self, **kwargs):
        context = super(DetalleCliente, self).get_context_data(**kwargs)
        context['cliente']= Cliente.objects.get(pk=self.kwargs['pk'])#get porque solo es para uno
        context['ordenes']= Orden.objects.filter(Cliente=self.kwargs['pk'])
        return context



class CrearCliente (CreateView):
    model = Cliente
    #fields Im gonna let users to fill
    fields =['Nombre','Apellido','Cedula','Direccion','Email', 'Gasto_acumulado', 'Telefono1', 'Telefono2']

    


class CrearClienteFromOrden(CrearCliente):
   
    success_url = reverse_lazy('index:Add-Orden')

    #def get_success_url():
    #    return reverse('index:Add-Orden')

    

class EditarCliente (UpdateView):
    model= Cliente
    #fields Im gonna let users to edit
    fields =['Telefono1', 'Telefono2']

class EliminarCliente (DeleteView):
    model= Cliente
    success_url= reverse_lazy("index:Home-Cliente")

    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)




#...............Equipo Views

class ListaEquipo(generic.ListView):
    #definir el template que utilizará y pasará el contexto
    template_name = 'index/equipos.html'
    #se define el nombre del contexto
    context_object_name='all_equipos'

    # @Override devuelve los objetos que serán renderizados
    def get_queryset(self):
        return Equipo.objects.all()
        
    def get_context_data(self, **kwargs):
        context = super(ListaEquipo, self).get_context_data(**kwargs)
        context['ultimos'] = Equipo.objects.order_by('-Fecha_creacion')[:5]
        return context

    

class CrearEquipo (CreateView):
    model = Equipo
    #fields Im gonna let users to fill
    fields =['Tipo','Marca','Modelo','Serial', 'Cliente']

class CrearEquipoFromOrden(CrearEquipo):
    
    success_url = reverse_lazy('index:Add-Orden')

class EliminarEquipo (DeleteView):
    model= Equipo
    success_url= reverse_lazy("index:Home-Equipo")

    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)


#............................TECNICO Views...........................

class ListaTecnico(generic.ListView):
    #definir el template que utilizará y pasará el contexto
    template_name = 'index/tecnicos.html'
    #se define el nombre del contexto
    context_object_name='all_tecnicos'

    # @Override devuelve los objetos que serán renderizados
    def get_queryset(self):
        return Tecnico.objects.all()
        
    

class CrearTecnico (CreateView):
    model = Tecnico
    #fields Im gonna let users to fill
    fields =['Nombre','Apellido']



class EliminarTecnico (DeleteView):
    model= Tecnico
    success_url= reverse_lazy("index:Home-Tecnico")

    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)


    